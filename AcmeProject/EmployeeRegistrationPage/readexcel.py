import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select




def readexcel(driver):
    wb = openpyxl.load_workbook("Datafile.xlsx")
    sheet = wb['Details']

    rowCount = wb.active.max_row
    colCount = wb.active.max_column


    for curr_row in range(2, rowCount+1):
        #####################Entering Employee details
        driver.find_element(By.ID, "FL:_ctl0:_ctl3").send_keys(sheet.cell(curr_row,1).value)       #("Isabel Britt")
        driver.find_element(By.ID, "FL:_ctl1:_ctl4").send_keys(sheet.cell(curr_row,2).value)       #("This is a test Employee Summary.")
        desig = driver.find_element(By.ID, "FL:_ctl3:_ctl3")
        desigvalue = Select(desig)
        desigvalue.select_by_visible_text(sheet.cell(curr_row,3).value)                            #("Management")
        driver.find_element(By.ID, "FL:_ctl4:_ctl3").send_keys(sheet.cell(curr_row,4).value)       #("$50,000.00")
        driver.find_element(By.ID, "FL_latTxt_5").send_keys(sheet.cell(curr_row,5).value)          #("34.833850°")
        driver.find_element(By.ID, "FL_longTxt_5").send_keys(sheet.cell(curr_row,6).value)         #("106.748580°")
        location = driver.find_element(By.ID, "FL:_ctl6:_ctl3")
        locationValue = Select(location)
        locationValue.select_by_visible_text(sheet.cell(curr_row,7).value)                         #("Headquarters")
        driver.find_element(By.ID, "FL:_ctl8:_ctl3").send_keys(sheet.cell(curr_row,8).value)       #("0020180104")
        if (sheet.cell(curr_row,10).value)=="Yes":
            driver.find_element(By.ID, "FL__ctl3_9").click()
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_12:Container']/div/div/div/span/input").send_keys(sheet.cell(curr_row,11).value)   #("47")
        lenM = driver.find_element(By.XPATH,
                                   "//body[1]/div[1]/div[1]/div[15]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/div[1]/div[1]/span[1]/select[1]")
        if (sheet.cell(curr_row,12).value)=="in" :
            lenMValue = Select(lenM)
            lenMValue.select_by_visible_text("in")
        elif (sheet.cell(curr_row,12).value) == "mm" :
            lenMValue = Select(lenM)
            lenMValue.select_by_visible_text("mm")
        elif (sheet.cell(curr_row,12).value) == "cm" :
            lenMValue = Select(lenM)
            lenMValue.select_by_visible_text("cm")

        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_13:Container']/div/div/div/span/input").send_keys(sheet.cell(curr_row,13).value)  #("21")
        widM = driver.find_element(By.XPATH,
                                   "//body[1]/div[1]/div[1]/div[15]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[3]/div[1]/div[1]/div[1]/span[1]/select[1]")
        if (sheet.cell(curr_row,14).value)=="in" :
            widMValue = Select(widM)
            widMValue.select_by_visible_text("in")
        elif (sheet.cell(curr_row,14).value) == "mm" :
            widMValue = Select(widM)
            widMValue.select_by_visible_text("mm")
        elif (sheet.cell(curr_row,14).value) == "cm" :
            widMValue = Select(widM)
            widMValue.select_by_visible_text("cm")

        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_14:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,15).value)    #("Brown")
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_16:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,17).value)    #("Ford")
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_17:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,18).value)    #("Taurus")
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_18:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,19).value)    #("2018")
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_19:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,20).value)    #("SEL")
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_20:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,21).value)    #("Black")
        driver.find_element(By.XPATH, "//*[@id='Row0.Field500_21:Container']/div/div/div/input").send_keys(sheet.cell(curr_row,22).value)    #("TEST-0001")
        driver.find_element(By.XPATH, "//tbody/tr[3]/td[2]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,24).value)          #("Ford")
        driver.find_element(By.XPATH, "//tbody/tr[3]/td[3]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,25).value)          #("F150")
        driver.find_element(By.XPATH, "//tbody/tr[3]/td[4]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,26).value)          #("2015")
        driver.find_element(By.XPATH, "//tbody/tr[3]/td[5]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,27).value)          #("XLT")
        driver.find_element(By.XPATH, "//tbody/tr[3]/td[6]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,28).value)          #("Red")
        driver.find_element(By.XPATH, "//tbody/tr[3]/td[7]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,29).value)          #("TEST-0002")
        driver.find_element(By.XPATH, "//tbody/tr[4]/td[2]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,31).value)          #("")
        driver.find_element(By.XPATH, "//tbody/tr[4]/td[3]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,32).value)          #("")
        driver.find_element(By.XPATH, "//tbody/tr[4]/td[4]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,33).value)          #("")
        driver.find_element(By.XPATH, "//tbody/tr[4]/td[5]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,34).value)          #("")
        driver.find_element(By.XPATH, "//tbody/tr[4]/td[6]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,35).value)          #("")
        driver.find_element(By.XPATH, "//tbody/tr[4]/td[7]/div[1]/div[1]/div[1]/input[1]").send_keys(sheet.cell(curr_row,36).value)          #("")








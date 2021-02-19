from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import openpyxl



def assertMethod(driver,userurl,empName):

    ############## getting data from excel
    wb = openpyxl.load_workbook("Datafile.xlsx")
    sheet = wb['Details']

    rowCount = wb.active.max_row
    colCount = wb.active.max_column



    for curr_row in range(2, rowCount + 1):
        if (sheet.cell(curr_row,1).value) == empName :
            ##############Verifing Employee details
            try:
                assert driver.find_element(By.ID,"Field.500_1:ValueContainer").text == (sheet.cell(curr_row,1).value) #"Isabel Britt"
            except:
                exceptionHandler("Name Error", "Employee name Mismatch", "Employee name should be -"+(sheet.cell(curr_row,1).value) )

            try:
                employeeSummary=driver.find_element(By.ID,"Field.500_2:ValueContainer").text
                assert employeeSummary.endswith(sheet.cell(curr_row,2).value)                #("This is a test Employee Summary.")
            except:
                exceptionHandler("Summery Error", "Employee Summery Mismatch",
                                 "Employee summery should be -" + (sheet.cell(curr_row, 2).value))

            try:
                assert driver.find_element(By.ID,"Field.500_7:ValueContainer").text == (sheet.cell(curr_row, 3).value)  #"Management"
            except:
                exceptionHandler("Department Error", "Department Mismatch",
                                 "Department should be -" + (sheet.cell(curr_row, 3).value))

            try:
                sal= driver.find_element(By.ID,"Field.500_6:ValueContainer").text
               # print("salary:",sal)
                exsal= sheet.cell(curr_row, 4).value
                #print('excel sal:',exsal  )
                exsal1= "${:,.2f}".format(exsal)
                #print(exsal1)
                assert sal == exsal1     #"$50,000.00"
            except:
                exceptionHandler("Salary Error", "Incorrect salary  : "+sal,
                                 "Salary should be  : " + str(sheet.cell(curr_row, 4).value))

            address= (sheet.cell(curr_row, 5).value)+", "+(sheet.cell(curr_row, 6).value)+" Map"
            try:
                loc= driver.find_element(By.ID,"Field.500_25:ValueContainer").text

                assert loc== str(address)   #34.833850°, 106.748580° Map"
            except:
                exceptionHandler("Address Error", "Address Mismatch : "+loc,
                                 "Address should be : " + address)

            try:
                assert driver.find_element(By.ID,"Field.500_8:ValueContainer").text == (sheet.cell(curr_row, 7).value) # "Headquarters"
            except:
                exceptionHandler("Work Location", "Work Location Mismatch",
                                 "Work Location should be -" + (sheet.cell(curr_row, 7).value))

            try:
                doj= driver.find_element(By.ID,"Field.500_3:ValueContainer").text
               # exdoj ="Jan 4, 2018"
                exdoj= sheet.cell(2,37).value
                print(doj)
                print(exdoj)
                assert doj == exdoj #"Jun 4, 2018"
            except:
                exceptionHandler("Date of Joining Error", "Date of Joining Mismatch : "+doj,
                                 "Date of Joining should be : "  + exdoj)

            try:
                assert driver.find_element(By.ID,"Field.500_4:ValueContainer").text == (sheet.cell(curr_row,10).value)   #"Yes"
            except:
                exceptionHandler("Active status Error", "Active status Mismatch",
                                 "Active status should be -" + (sheet.cell(curr_row, 10).value))
            lenvalue=str(sheet.cell(curr_row, 11).value)+(sheet.cell(curr_row, 12).value)
            try:
                assert driver.find_element(By.XPATH,"//*[@id='Row0.Field500_12:Container']/div/div/div").\
                           text == lenvalue         #"47in"
            except:
                exceptionHandler("Employee cubical need - lenght Error", "Lenght Mismatch",
                                 "Lenght should be -" + lenvalue)
            widvalue = str(sheet.cell(curr_row, 13).value) + (sheet.cell(curr_row, 14).value)

            try:
                assert driver.find_element(By.XPATH,"//*[@id='Row0.Field500_13:Container']/div/div/div").text == widvalue   #"21in"
            except :
                exceptionHandler("Employee cubical need - Width Error", "Width Mismatch",
                                 "Width should be -" + widvalue)
            try:
                assert driver.find_element(By.XPATH,"//*[@id='Row0.Field500_14:Container']/div/div/div").text == (sheet.cell(curr_row, 15).value)   # "Brown"
            except:
                exceptionHandler("Color Error", "Color Mismatch",
                                 "Color should be -" + (sheet.cell(curr_row, 15).value))

            ################# Car 1 Assertion
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 17).value)   # "Ford"
            except :
                exceptionHandler("Car1 Brand Error", "Brand Mismatch",
                                 "Brand should be -" + (sheet.cell(curr_row, 17).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[3]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 18).value)  #"Taurus"
            except:
                exceptionHandler("Car1 Model Error", "Model Mismatch",
                                 "Model should be -" + (sheet.cell(curr_row, 18).value))

            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[4]/div[1]/div[1]/div[1]").text == str(sheet.cell(curr_row, 19).value)  #"2018"
            except:
                exceptionHandler("Car1 Model Year Error", "Model Year Mismatch",
                                 "Model Year should be -" + str(sheet.cell(curr_row, 19).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[5]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 20).value) #"SEL"
            except:
                exceptionHandler("Car1 Trim Error", "Trim Mismatch",
                                 "Trim should be -" + (sheet.cell(curr_row, 20).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[6]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 21).value)  # "Black"
            except:
                exceptionHandler("Car1 Colour Error", "Colour Mismatch",
                                 "Colour should be -" + (sheet.cell(curr_row, 21).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[2]/td[7]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 22).value)  #"TEST-0001"
            except:
                exceptionHandler("Car1 License Plate Error", "License Plate Mismatch",
                                 "License Plate should be -" + (sheet.cell(curr_row, 22).value))

            ################## Car2 Assertion
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[2]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 24).value) #"Ford"
            except:
                exceptionHandler("Car2 Brand Error", "Brand Mismatch",
                                 "Brand should be -" + (sheet.cell(curr_row, 24).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[3]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 25).value)  #"F150"
            except:
                exceptionHandler("Car2 Model Error", "Model Mismatch",
                                 "Model should be -" + (sheet.cell(curr_row, 25).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/div[1]/div[1]").text == str(sheet.cell(curr_row, 26).value)  #"2015"
            except:
                exceptionHandler("Car2 Model Year Error", "Model Year Mismatch",
                                 "Model Year should be -" + str(sheet.cell(curr_row, 26).value))

            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[5]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 27).value)  #"XLT"
            except:
                exceptionHandler("Car2 Trim Error", "Trim Mismatch",
                                 "Trim should be -" + (sheet.cell(curr_row, 27).value))
            try:
                assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[6]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 28).value)  #"Red"
            except:
                exceptionHandler("Car2 colour Error", "Colour Mismatch",
                                 "Colour should be -" + (sheet.cell(curr_row, 28).value))
            try:
                lp= driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[7]/div[1]/div[1]/div[1]").text
                assert str(lp).lower() == str(sheet.cell(curr_row, 29).value).lower()  # "Test-0002"
            except:
                exceptionHandler("Car2 License Plate Error", "License plate Mismatch : "+lp,
                                 "License plate  should be : " + (sheet.cell(curr_row, 29).value))

            ################ Car 3 details
            car3brand= (sheet.cell(curr_row, 31).value)
            if bool(car3brand)== "True":
                try:
                    assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[2]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 31).value) #""
                except:
                    exceptionHandler("Car1 Brand Error", "Brand Mismatch",
                                     "Brand should be -" + (sheet.cell(curr_row, 31).value))
                try:
                    assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[3]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 32).value)  #""
                except:
                    exceptionHandler("Car3 Model Error", "Model Mismatch",
                                     "Model should be -" + (sheet.cell(curr_row, 32).value))
                try:
                    assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 33).value)  #""
                except:
                    exceptionHandler("Car3 Model Year Error", "Model Year Mismatch",
                                     "Model Year should be -" + str(sheet.cell(curr_row, 33).value))

                try:
                    assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[5]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 34).value)  #""
                except:
                    exceptionHandler("Car3 Trim Error", "Trim Mismatch",
                                     "Trim should be -" + (sheet.cell(curr_row, 34).value))
                try:
                    assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[6]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 35).value)  #""
                except:
                    exceptionHandler("Car3 colour Error", "Colour Mismatch",
                                     "Colour should be -" + (sheet.cell(curr_row, 35).value))
                try:
                    assert driver.find_element(By.XPATH,"//body[1]/div[1]/div[2]/div[17]/span[2]/div[1]/table[1]/tbody[1]/tr[3]/td[7]/div[1]/div[1]/div[1]").text == (sheet.cell(curr_row, 36).value)  # ""
                except:
                    exceptionHandler("Car3 License Plate Error", "License plate Mismatch",
                                     "License plate  should be -" + (sheet.cell(curr_row, 36).value))



def exceptionHandler(header,description,expected) :
    with open("Issue_Log.txt", "a") as file:
        file.write("\n")
        file.write("a.Issue Heading          : "+ header)
        file.write("\n")
        file.write("b. Issue Description     : "+ description)
        file.write("\n")
        file.write("c. Expected Behaviour    : "+ expected)
        file.write("\n")
        file.write("-----------------------------------------------------------")

